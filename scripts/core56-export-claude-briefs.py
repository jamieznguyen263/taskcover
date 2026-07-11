from __future__ import annotations

import argparse
import csv
import io
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as error:
    raise SystemExit(
        "openpyxl is required. Run this with the Codex bundled Python runtime or install openpyxl in the active Python environment."
    ) from error


DEFAULT_WORKBOOK = Path(
    r"C:\Users\Gamelap\OneDrive\Desktop\Taskcover.com\Taskcover_Core_56_Final_Outlines_Top10_Readiness_9_5.xlsx"
)
DEFAULT_OUTPUT = Path("docs/core56-claude-batches")

BATCHES = [
    ("batch-01-reconcile", ["TC-001", "TC-006"], "Already Started / Reconcile"),
    ("batch-02-ai-search-core", ["TC-007", "TC-002", "TC-008", "TC-009", "TC-010"], "Wave 1 AI Search Core"),
    ("batch-03-wave1-search-buying-strategy", ["TC-003", "TC-004", "TC-005", "TC-011", "TC-012", "TC-013"], "Wave 1 Search Buying And Strategy"),
    ("batch-04-measurement-technical-seo", ["TC-014", "TC-016", "TC-017", "TC-018", "TC-031", "TC-032", "TC-033", "TC-034"], "Measurement And Technical SEO"),
    ("batch-05-content-topical-authority", ["TC-019", "TC-020", "TC-035", "TC-036", "TC-037"], "Content, Topical Authority, Internal Links"),
    ("batch-06-digital-pr-authority", ["TC-021", "TC-038", "TC-039", "TC-040"], "Digital PR And Authority"),
    ("batch-07-local-franchise-international", ["TC-022", "TC-023", "TC-024", "TC-044", "TC-045", "TC-046", "TC-047"], "Local, Franchise, International SEO"),
    ("batch-08-industry-seo", ["TC-025", "TC-026", "TC-027", "TC-028", "TC-029", "TC-030"], "Industry SEO"),
    ("batch-09-ecommerce-saas", ["TC-048", "TC-049", "TC-050", "TC-051", "TC-052"], "Ecommerce And SaaS"),
    ("batch-10-ppc-enablement-reporting-benchmark", ["TC-053", "TC-054", "TC-055", "TC-056", "TC-015"], "PPC, Enablement, Reporting, Benchmark"),
    ("batch-11-ai-search-entity-prompt-scorecard", ["TC-041", "TC-042", "TC-043"], "AI Search Entity, Prompt Research, Scorecard"),
]

BRIEF_FIELDS = [
    "H1 / Final title",
    "Suggested URL slug",
    "Meta title",
    "Meta description",
    "Primary keyword",
    "Search intent",
    "Target market",
    "Target words",
    "Opening answer requirement",
    "Mandatory FAQs",
    "Unique evidence / asset",
    "Source keys",
    "Author / reviewer requirement",
    "Regional localization instructions",
    "Internal links",
    "Recommended schema",
    "Visual / data plan",
    "Conversion CTA",
    "Forbidden claims",
    "Update cycle",
    "Brief status",
]

MASTER_FIELDS = [
    "Priority",
    "Cluster",
    "Role",
    "Wave",
    "Primary money page",
    "Supporting pages",
    "Format",
    "Unique angle / information gain",
    "Original asset",
    "Target words",
    "Provisional outline score",
    "Readiness",
    "Primary ranking risk",
    "Update cycle",
    "Status",
]


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Core 56 workbook briefs into Claude-ready batch Markdown files.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--validate-only", action="store_true", help="Validate workbook and generated batch coverage without writing files.")
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    workbook = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    master = keyed_rows(workbook["Core 56 Master"], "Article ID")
    briefs = keyed_rows(workbook["Final Writer Briefs"], "Article ID")
    outlines = outline_rows(workbook["Outline Sections"])
    sources = keyed_rows(workbook["Sources & Governance"], "Source key")

    validation = validate_batches(master, briefs, outlines)
    if not validation["passed"]:
        raise SystemExit(json.dumps(validation, indent=2))

    if args.validate_only:
        print(json.dumps(validation, indent=2))
        return

    args.out.mkdir(parents=True, exist_ok=True)
    for batch_slug, article_ids, batch_title in BATCHES:
        path = args.out / f"{batch_slug}.md"
        path.write_text(render_batch(batch_title, article_ids, master, briefs, outlines, sources), encoding="utf-8")

    index_path = args.out / "README.md"
    index_path.write_text(render_index(), encoding="utf-8")
    manifest_path = args.out / "manifest.json"
    manifest = render_manifest(master, briefs, outlines)
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    tracker_path = args.out / "publication-tracker.csv"
    tracker_path.write_text(render_publication_tracker(manifest), encoding="utf-8")
    print(f"Exported {len(BATCHES)} batch brief files, manifest.json, and publication-tracker.csv to {args.out}")


def keyed_rows(sheet: Any, key_field: str) -> dict[str, dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    key_index = headers.index(key_field)
    output: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        if not row or row[key_index] is None:
            continue
        key = str(row[key_index]).strip()
        output[key] = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
    return output


def outline_rows(sheet: Any) -> dict[str, list[dict[str, Any]]]:
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    article_index = headers.index("Article ID")
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows[1:]:
        if not row or row[article_index] is None:
            continue
        item = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        result[str(row[article_index]).strip()].append(item)
    for article_rows in result.values():
        article_rows.sort(key=lambda item: int(item.get("Section order") or 0))
    return result


def render_index() -> str:
    lines = [
        "# Core 56 Claude Batch Briefs",
        "",
        "These files are generated from the workbook writer briefs and outline sections.",
        "Use them with `docs/core56-claude-master-prompt.vi.md`.",
        "",
        "Machine-readable tracking lives in `manifest.json`.",
        "Publication workflow tracking starts in `publication-tracker.csv`.",
        "",
        "Validate coverage before sending work to Claude:",
        "",
        "```powershell",
        "& 'C:\\Users\\Gamelap\\.cache\\codex-runtimes\\codex-primary-runtime\\dependencies\\python\\python.exe' scripts/core56-export-claude-briefs.py --validate-only",
        "```",
        "",
        "Regenerate this folder from the workbook:",
        "",
        "```powershell",
        "& 'C:\\Users\\Gamelap\\.cache\\codex-runtimes\\codex-primary-runtime\\dependencies\\python\\python.exe' scripts/core56-export-claude-briefs.py",
        "```",
        "",
        "| Batch | Articles | File |",
        "|---|---|---|",
    ]
    for batch_slug, article_ids, batch_title in BATCHES:
        lines.append(f"| {batch_title} | {', '.join(article_ids)} | `{batch_slug}.md` |")
    lines.append("")
    return "\n".join(lines)


def validate_batches(
    master: dict[str, dict[str, Any]],
    briefs: dict[str, dict[str, Any]],
    outlines: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    batch_ids = [article_id for _, article_ids, _ in BATCHES for article_id in article_ids]
    workbook_ids = sorted(master.keys())
    duplicate_ids = sorted(article_id for article_id, count in count_values(batch_ids).items() if count > 1)
    missing_from_workbook = sorted(set(batch_ids) - set(master.keys()) | (set(batch_ids) - set(briefs.keys())))
    missing_from_batches = sorted(set(workbook_ids) - set(batch_ids))
    extra_in_batches = sorted(set(batch_ids) - set(workbook_ids))
    empty_outlines = sorted(article_id for article_id in batch_ids if not outlines.get(article_id))
    missing_required_fields = [
        {"articleId": article_id, "field": field}
        for article_id in batch_ids
        for field in ["H1 / Final title", "Suggested URL slug", "Meta title", "Meta description", "Primary keyword", "Internal links", "Forbidden claims"]
        if not text(briefs.get(article_id, {}).get(field))
    ]
    passed = not (duplicate_ids or missing_from_workbook or missing_from_batches or extra_in_batches or empty_outlines or missing_required_fields)
    return {
        "passed": passed,
        "batchCount": len(BATCHES),
        "expectedArticleCount": len(workbook_ids),
        "batchedArticleCount": len(batch_ids),
        "duplicateIds": duplicate_ids,
        "missingFromWorkbook": missing_from_workbook,
        "missingFromBatches": missing_from_batches,
        "extraInBatches": extra_in_batches,
        "emptyOutlines": empty_outlines,
        "missingRequiredFields": missing_required_fields,
    }


def render_manifest(
    master: dict[str, dict[str, Any]],
    briefs: dict[str, dict[str, Any]],
    outlines: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    return {
        "sourceWorkbook": str(DEFAULT_WORKBOOK),
        "batchCount": len(BATCHES),
        "articleCount": sum(len(article_ids) for _, article_ids, _ in BATCHES),
        "batches": [
            {
                "slug": batch_slug,
                "title": batch_title,
                "file": f"{batch_slug}.md",
                "articleIds": article_ids,
            }
            for batch_slug, article_ids, batch_title in BATCHES
        ],
        "articles": [
            {
                "articleId": article_id,
                "batchSlug": batch_slug,
                "batchTitle": batch_title,
                "priority": master[article_id].get("Priority"),
                "wave": master[article_id].get("Wave"),
                "role": master[article_id].get("Role"),
                "cluster": master[article_id].get("Cluster"),
                "title": briefs[article_id].get("H1 / Final title"),
                "slug": normalize_slug(briefs[article_id].get("Suggested URL slug")),
                "primaryKeyword": briefs[article_id].get("Primary keyword"),
                "moneyPage": master[article_id].get("Primary money page"),
                "targetWords": briefs[article_id].get("Target words") or master[article_id].get("Target words"),
                "sourceKeys": sorted(split_source_keys(briefs[article_id].get("Source keys"))),
                "outlineSectionCount": len(outlines.get(article_id, [])),
                "briefStatus": briefs[article_id].get("Brief status"),
                "readiness": master[article_id].get("Readiness"),
                "primaryRankingRisk": master[article_id].get("Primary ranking risk"),
            }
            for batch_slug, article_ids, batch_title in BATCHES
            for article_id in article_ids
        ],
    }


def render_publication_tracker(manifest: dict[str, Any]) -> str:
    buffer = io.StringIO()
    fieldnames = [
        "article_id",
        "batch_slug",
        "batch_title",
        "priority",
        "wave",
        "role",
        "cluster",
        "title",
        "slug",
        "primary_keyword",
        "money_page",
        "claude_brief_file",
        "claude_output_status",
        "validation_status",
        "converted_backfill_file",
        "import_status",
        "publish_status",
        "live_url",
        "post_publish_qa_status",
        "notes",
    ]
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    batch_file_by_slug = {batch["slug"]: batch["file"] for batch in manifest["batches"]}
    batch_title_by_slug = {batch["slug"]: batch["title"] for batch in manifest["batches"]}
    for article in manifest["articles"]:
        writer.writerow(
            {
                "article_id": article["articleId"],
                "batch_slug": article["batchSlug"],
                "batch_title": batch_title_by_slug.get(article["batchSlug"], article["batchTitle"]),
                "priority": article["priority"],
                "wave": article["wave"],
                "role": article["role"],
                "cluster": article["cluster"],
                "title": article["title"],
                "slug": article["slug"],
                "primary_keyword": article["primaryKeyword"],
                "money_page": article["moneyPage"],
                "claude_brief_file": batch_file_by_slug.get(article["batchSlug"], ""),
                "claude_output_status": "not-received",
                "validation_status": "not-run",
                "converted_backfill_file": "",
                "import_status": "not-imported",
                "publish_status": "not-published",
                "live_url": "",
                "post_publish_qa_status": "not-run",
                "notes": article["primaryRankingRisk"],
            }
        )
    return buffer.getvalue()


def render_batch(
    batch_title: str,
    article_ids: list[str],
    master: dict[str, dict[str, Any]],
    briefs: dict[str, dict[str, Any]],
    outlines: dict[str, list[dict[str, Any]]],
    sources: dict[str, dict[str, Any]],
) -> str:
    lines = [
        f"# Core 56 Claude Batch Brief - {batch_title}",
        "",
        "Use this file together with `docs/core56-claude-master-prompt.vi.md`.",
        "Write complete English article packages only for the article IDs in this batch.",
        "",
        "## Batch Articles",
        "",
        "| ID | Title | Primary keyword | Money page |",
        "|---|---|---|---|",
    ]
    for article_id in article_ids:
        brief = briefs[article_id]
        master_row = master[article_id]
        lines.append(
            f"| {article_id} | {md_inline(brief.get('H1 / Final title'))} | {md_inline(brief.get('Primary keyword'))} | {md_inline(master_row.get('Primary money page'))} |"
        )

    lines.extend(["", "## Article Briefs", ""])
    used_source_keys: set[str] = set()
    for article_id in article_ids:
        brief = briefs[article_id]
        master_row = master[article_id]
        used_source_keys.update(split_source_keys(brief.get("Source keys")))
        lines.extend(render_article(article_id, master_row, brief, outlines.get(article_id, [])))

    lines.extend(render_sources(used_source_keys, sources))
    return "\n".join(lines)


def render_article(article_id: str, master_row: dict[str, Any], brief: dict[str, Any], outline: list[dict[str, Any]]) -> list[str]:
    lines = [f"### {article_id} - {text(brief.get('H1 / Final title'))}", ""]
    lines.extend(["#### Master row", ""])
    for field in MASTER_FIELDS:
        append_field(lines, field, master_row.get(field))

    lines.extend(["", "#### Writer brief", ""])
    for field in BRIEF_FIELDS:
        append_field(lines, field, normalize_slug(brief.get(field)) if field == "Suggested URL slug" else brief.get(field))

    lines.extend(["", "#### Required outline sections", ""])
    if not outline:
        lines.append("- No outline sections found.")
    for item in outline:
        level = text(item.get("Heading level"))
        heading = text(item.get("Heading / block"))
        order = text(item.get("Section order"))
        lines.append(f"{order}. **{md_inline(level)} - {md_inline(heading)}**")
        append_nested(lines, "Coverage", item.get("Coverage instructions"))
        append_nested(lines, "Evidence/output", item.get("Evidence / output required"))
        append_nested(lines, "Market handling", item.get("Market handling"))
    lines.append("")
    return lines


def render_sources(source_keys: set[str], sources: dict[str, dict[str, Any]]) -> list[str]:
    lines = ["## Source Keys For This Batch", ""]
    if not source_keys:
        lines.append("- No source keys listed.")
        return lines
    for key in sorted(source_keys):
        row = sources.get(key)
        if not row:
            lines.append(f"- `{key}`: missing from Sources & Governance sheet; verify manually.")
            continue
        lines.append(f"- `{key}`: {text(row.get('Source'))} - {text(row.get('URL'))}")
        append_nested(lines, "Used for", row.get("Used for"))
        append_nested(lines, "Editorial instruction", row.get("Editorial instruction"))
    lines.append("")
    return lines


def append_field(lines: list[str], label: str, value: Any) -> None:
    cleaned = text(value)
    if cleaned:
        lines.append(f"- **{label}:** {cleaned}")


def append_nested(lines: list[str], label: str, value: Any) -> None:
    cleaned = text(value)
    if cleaned:
        lines.append(f"  - {label}: {cleaned}")


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def md_inline(value: Any) -> str:
    return text(value).replace("|", "/").replace("\n", " ")


def normalize_slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text(value).lower()).strip("-")


def split_source_keys(value: Any) -> set[str]:
    if value is None:
        return set()
    return {part.strip() for part in re.split(r"[|,;]", str(value)) if part.strip()}


def count_values(values: list[str]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for value in values:
        counts[value] += 1
    return counts


if __name__ == "__main__":
    main()
